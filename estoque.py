import customtkinter as ctk
from CTkMessagebox import CTkMessagebox
from openpyxl import load_workbook, Workbook
import json
import os

arquivo_estoque = 'banco.json'
arquivo_excel = 'estoque.xlsx'

def carregar_produtos():
    if os.path.exists(arquivo_estoque):
        with open(arquivo_estoque, 'r') as a:
            try:
                return json.load(a)
            except json.JSONDecodeError:
                return []
    return []

def salvar_produto(produtos):
    with open(arquivo_estoque, 'w') as a:
        json.dump(produtos, a, indent=4)

def criar_arquivo_excel():
    if not os.path.exists(arquivo_excel):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Estoque'
        ws.append(['Código', 'Nome', 'Categoria', 'Quantidade', 'Preço'])
        wb.save(arquivo_excel)

def limpar_campos():
    entrada_codigo.delete(0, 'end')
    entrada_nome.delete(0, 'end')
    entrada_categoria.delete(0, 'end')
    entrada_quantidade.delete(0, 'end')
    entrada_preco.delete(0, 'end')

def adicionar_produto():
    criar_arquivo_excel()
    codigo = entrada_codigo.get()
    nome = entrada_nome.get()
    categoria = entrada_categoria.get()
    quantidade = entrada_quantidade.get()
    preco = entrada_preco.get()

    if not all([codigo, nome, categoria, quantidade, preco]):
        CTkMessagebox(title='ERRO', message='Preencha todos os campos!', icon='cancel')
        return

    try:
        quantidade = int(quantidade)
        preco = entrada_preco.get().replace(',', '.')
        preco = float(preco)
    except ValueError:
        CTkMessagebox(title='ERRO', message='Quantidade e preço devem ser numéricos!')
        return

    produtos = carregar_produtos()
    produtos.append({
        'codigo': codigo,
        'nome': nome,
        'categoria': categoria,
        'quantidade': quantidade,
        'preco': preco
    })
    salvar_produto(produtos)

    wb = load_workbook(arquivo_excel)
    ws = wb.active
    ws.append([codigo, nome, categoria, quantidade, preco])
    wb.save(arquivo_excel)

    CTkMessagebox(title='Sucesso', message='Produto adicionado com sucesso!', icon='check')
    limpar_campos()
        
def remover_produto():
    pass
    
janela = ctk.CTk()
janela.title('Gerenciador de estoque')
janela.geometry('750x450')

frame_form = ctk.CTkFrame(janela)
frame_form.pack(pady=10)

codigo_label = ctk.CTkLabel(frame_form, text='Código:')
codigo_label.grid(row=0, column=0)
entrada_codigo = ctk.CTkEntry(frame_form, width=200)
entrada_codigo.grid(row=0, column=1)

nome_label = ctk.CTkLabel(frame_form, text='Nome:')
nome_label.grid(row=0, column=2)
entrada_nome = ctk.CTkEntry(frame_form, width=200)
entrada_nome.grid(row=0, column=3)

categoria_label = ctk.CTkLabel(frame_form, text='Categoria:')
categoria_label.grid(row=1, column=0)
entrada_categoria = ctk.CTkEntry(frame_form, width=200)
entrada_categoria.grid(row=1, column=1)

quantidade_label = ctk.CTkLabel(frame_form, text='Quantidade:')
quantidade_label.grid(row=1, column=2)
entrada_quantidade = ctk.CTkEntry(frame_form, width=200)
entrada_quantidade.grid(row=1, column=3)

preco_label = ctk.CTkLabel(frame_form, text='Preço:')
preco_label.grid(row=2, column=0)
entrada_preco = ctk.CTkEntry(frame_form, width=200)
entrada_preco.grid(row=2, column=1)

adicionar_botao = ctk.CTkButton(frame_form, text='Adicionar', fg_color='green'
                                , command=adicionar_produto, width=50)
adicionar_botao.grid(row=3, column=0, pady=10, padx=10)

remover_botao = ctk.CTkButton(frame_form, text='Remover', fg_color='red'
                                , command=remover_produto, width=50)
remover_botao.grid(row=3, column=1, pady=10, padx=10)

janela.mainloop()